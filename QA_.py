
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import time
from tqdm import tqdm
from google.colab import drive

# Mount Google Drive (if your files are stored there)
# Uncomment the following lines if you need to access files from Google Drive
# drive.mount('/content/drive')

def create_key(df):
    df['Key'] = df['ChemicalID'].astype(str) + '_' + df['PartNumber'].astype(str)
    return df

def validate_rows_count(df):
    actual_counts = df['Key'].value_counts().reset_index()
    actual_counts.columns = ['Key', 'ActualRowsCount']

    df = df.merge(actual_counts, on='Key')
    df['RowsCountGap'] = df['RowsCount '] - df['ActualRowsCount']

    df['Automated QA Comment'] = df.apply(
        lambda x: (x['Automated QA Comment'] + ' | ' if x['Automated QA Comment'] else '') + 'Rows count mismatch'
        if x['RowsCountGap'] != 0 else x['Automated QA Comment'], axis=1
    )

    return df

def check_fmd_revision_flag(df):
    df['Automated QA Comment'] = df.apply(
        lambda x: (x['Automated QA Comment'] + ' | ' if x['Automated QA Comment'] else '') + 'FMDRevFlag is Not Latest'
        if x['FMDRevFlag'] == 'Not Latest' else x['Automated QA Comment'], axis=1
    )
    return df

def check_homogeneous_material_mass_variation(df):
    df['HomogeneousMaterialName'] = df['HomogeneousMaterialName'].str.lower()
    grouped = df.groupby(['Key', 'HomogeneousMaterialName'])

    for (key, material_name), group in grouped:
        unique_masses = group['HomogeneousMaterialMass '].nunique()

        if unique_masses > 1:
            df.loc[(df['Key'] == key) & (df['HomogeneousMaterialName'] == material_name), 'Automated QA Comment'] = df.loc[(df['Key'] == key) & (df['HomogeneousMaterialName'] == material_name), 'Automated QA Comment'].apply(
                lambda x: x + ' | Multiple masses for the same homogeneous material' if x else 'Multiple masses for the same homogeneous material'
            )

    return df

def check_homogeneous_material_mass(df):
    df['HomogeneousMaterialName'] = df['HomogeneousMaterialName'].str.lower()
    df['CalculatedMass'] = df.groupby(['Key', 'HomogeneousMaterialName'])['Mass '].transform('sum')
    df['Homogeneous Mass Gap'] = df['CalculatedMass'] - df['HomogeneousMaterialMass ']
    df['MassMismatch'] = df['Homogeneous Mass Gap'].abs() >= 1

    df['Automated QA Comment'] = df.apply(
        lambda x: (x['Automated QA Comment'] + ' | ' if x['Automated QA Comment'] else '') + 'Fail: Mass mismatch'
        if x['MassMismatch'] else x['Automated QA Comment'], axis=1
    )

    return df

def check_substance_homogeneous_material_percentage(df):
    df['homogeneousPercentageSum'] = df.groupby(['Key', 'HomogeneousMaterialName'])['SubstanceHomogeneousMaterialPercentage '].transform('sum')
    df['PercentageMatch'] = (df['homogeneousPercentageSum'] >= 99.9) & (df['homogeneousPercentageSum'] <= 100.10)
    df['PercentageMatchComment'] = df.apply(lambda x: 'Fail: homogeneousPercentage sum != 100' if not x['PercentageMatch'] else '', axis=1)
    df['Automated QA Comment'] = df.apply(lambda x: x['Automated QA Comment'] + ' | ' + x['PercentageMatchComment'] if x['PercentageMatchComment'] else x['Automated QA Comment'], axis=1)
    return df

def check_substance_homogeneous_material_ppm(df):
    df['homogeneousPPMSum'] = df.groupby(['Key', 'HomogeneousMaterialName'])['SubstanceHomogeneousMaterialPercentagePPM '].transform('sum')
    df['PPMMatch'] = (df['homogeneousPPMSum'] >= 999000.0) & (df['homogeneousPPMSum'] <= 1001000.0)
    df['PPMMatchComment'] = df.apply(lambda x: 'Fail: homogeneousPPM sum != 1000000' if not x['PPMMatch'] else '', axis=1)
    df['Automated QA Comment'] = df.apply(lambda x: x['Automated QA Comment'] + ' | ' + x['PPMMatchComment'] if x['PPMMatchComment'] else x['Automated QA Comment'], axis=1)
    return df

def check_substance_component_level_percentage(df):
    df['ComponentPercentageSum'] = df.groupby('Key')['SubstanceComponentLevelPercentage '].transform('sum')
    df['ComponentPercentageMatch'] = (df['ComponentPercentageSum'] >= 99.0) & (df['ComponentPercentageSum'] <= 101.0)
    df['ComponentPercentageMatchComment'] = df.apply(lambda x: 'Fail: Component level percentage sum != 100' if not x['ComponentPercentageMatch'] else '', axis=1)
    df['Automated QA Comment'] = df.apply(lambda x: x['Automated QA Comment'] + ' | ' + x['ComponentPercentageMatchComment'] if x['ComponentPercentageMatchComment'] else x['Automated QA Comment'], axis=1)
    return df

def check_substance_component_level_ppm(df):
    df['ComponentPPMSum'] = df.groupby('Key')['SubstanceComponentLevelPPM '].transform('sum')
    df['ComponentPPMMatch'] = (df['ComponentPPMSum'] >= 990000.0) & (df['ComponentPPMSum'] <= 1010000.0)
    df['ComponentPPMMatchComment'] = df.apply(lambda x: 'Fail: Component level PPM sum != 1000000' if not x['ComponentPPMMatch'] else '', axis=1)
    df['Automated QA Comment'] = df.apply(lambda x: x['Automated QA Comment'] + ' | ' + x['ComponentPPMMatchComment'] if x['ComponentPPMMatchComment'] else x['Automated QA Comment'], axis=1)
    return df

def calculate_gap_and_comment(df):
    # Calculate the absolute gap
    gap = (df['TotalComponentMassProfile '] - df['TotalComponentMassSummation ']).abs()

    # Calculate the gap as a percentage of the TotalComponentMassProfile
    gap_percentage = (gap / df['TotalComponentMassProfile ']) * 100

    # Add comment to the existing 'Automated QA Comment' column if gap percentage >= 50%
    df['Automated QA Comment'] = df.apply(
        lambda x: (x['Automated QA Comment'] + ' | ' if x['Automated QA Comment'] else '') + 'Total VS Summation Gap is more than 50%'
        if (x['TotalComponentMassProfile '] != 0 and (abs(x['TotalComponentMassProfile '] - x['TotalComponentMassSummation '])/x['TotalComponentMassProfile '] * 100) >= 50) else x['Automated QA Comment'], axis=1
    )

    return df

def check_total_component_mass_summation(df):
    unique_keys = df['Key'].unique()
    for key in tqdm(unique_keys, desc="Checking Total Component Mass Summation"):
        group = df[df['Key'] == key]
        total_mass_sum = round(group['Mass '].sum(), 4)  # Rounding to 4 decimal places
        total_component_mass_summation = group['TotalComponentMassSummation '].iloc[0]

        if total_mass_sum != total_component_mass_summation:
            df.loc[df['Key'] == key, 'Automated QA Comment'] = df.loc[df['Key'] == key, 'Automated QA Comment'].apply(
                lambda x: x + ' | Software issue' if pd.notnull(x) and x else 'Software issue'
            )
    return df

def clear_worksheet_but_keep_header(worksheet):
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.value = None

def run_all_checks(file_path):
    start_time = time.time()

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    data = worksheet.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)

    # Ensure 'Automated QA Comment' column exists
    if 'Automated QA Comment' not in df.columns:
        df['Automated QA Comment'] = ''

    df = create_key(df)
    df = check_fmd_revision_flag(df)
    df = check_homogeneous_material_mass_variation(df)
    df = validate_rows_count(df)
    df = check_homogeneous_material_mass(df)
    df = check_substance_homogeneous_material_percentage(df)
    df = check_substance_homogeneous_material_ppm(df)
    df = check_substance_component_level_percentage(df)
    df = check_substance_component_level_ppm(df)
    df = calculate_gap_and_comment(df)

    # Progress tracking using tqdm
    df = check_total_component_mass_summation(df)

    # Select only the required columns
    added_columns = ['RowsCountGap', 'Homogeneous Mass Gap', 'homogeneousPercentageSum', 'homogeneousPPMSum', 'ComponentPercentageSum', 'ComponentPPMSum', 'Automated QA Comment']
    df = df[list(columns) + added_columns]  # Append added columns to original columns

    clear_worksheet_but_keep_header(worksheet)

    # Write headers explicitly and set font color to red for added columns
    for c_idx, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=c_idx, value=column)
        if column in added_columns:
            cell.font = Font(color="FF0000")  # Set font color to red for added columns

    # Write data starting from the second row
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Define output path
    if file_path.endswith('.xlsx'):
        output_path = file_path.replace('.xlsx', '_checked.xlsx')
    elif file_path.endswith('.xls'):
        output_path = file_path.replace('.xls', '_checked.xlsx')  # Save as .xlsx for consistency
    else:
        output_path = file_path + '_checked.xlsx'

    workbook.save(output_path)

    end_time = time.time()
    elapsed_time = end_time - start_time
    mins, secs = divmod(elapsed_time, 60)
    print(f"Process Complete. Processed file saved to {output_path}")
    print(f"Elapsed Time: {int(mins):02}:{int(secs):02}")

    return output_path

# ----------------------------
# Specify the file path below
# ----------------------------

# Example 1: If your file is in the current Colab environment
# file_path = '/content/your_file.xlsx'

# Example 2: If your file is stored in Google Drive
# Make sure to mount Google Drive first (uncomment drive.mount above)
# file_path = '/content/drive/MyDrive/path_to_your_file/your_file.xlsx'

# Replace the below path with your actual file path
file_path = '/content/qwa-9157bab7-17d8-40ca-934c-a45d395ef4dd.xlsx'  # <-- Update this path

# ----------------------------
# Run the QA Checks
# ----------------------------

# Check if the file exists
import os

if os.path.exists(file_path):
    print(f"Processing file: {file_path}")
    output_file = run_all_checks(file_path)
    print(f"Processed file is saved at: {output_file}")

    # Optional: If you want to download the file to your local machine
    from google.colab import files
    files.download(output_file)
else:
    print(f"File not found: {file_path}")
    print("Please check the file path and ensure the file exists.")


